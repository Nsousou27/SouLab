import os
import streamlit as st
import pandas as pd
import io
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="SouLab", layout="wide", page_icon="🧬")

css_path = os.path.join(os.path.dirname(__file__), "styles.css")
with open(css_path) as f:
    st.markdown("<style>" + f.read() + "</style>", unsafe_allow_html=True)

rows = list("ABCDEFGH")
cols = list(range(1, 13))
default_std_series = ["2000*", "1000*", "500*", "250*", "125*", "62.5*", "31.25*", "0*"]
EXPECTED_WELL_COLS = [f"{r}{c}" for r in "ABCDEFGH" for c in range(1, 13)]

def instr(text):
    html = (
        "<div style='background:rgba(58,134,255,0.07);border-left:3px solid #3a86ff;"
        "border-radius:0 8px 8px 0;padding:0.85rem 1.1rem;margin-bottom:1.1rem;'>"
        "<p style='margin:0;font-size:0.9rem;color:#000000;line-height:1.75;'>"
        + text +
        "</p></div>"
    )
    st.markdown(html, unsafe_allow_html=True)

def section_label(text):
    html = (
        "<div style='display:flex;align-items:center;gap:1rem;margin:2rem 0 1rem;'>"
        "<div style='flex:1;height:1px;background:linear-gradient(90deg,rgba(58,134,255,0.4),transparent);'></div>"
        "<div style='font-family:monospace;font-size:0.75rem;color:#3a86ff;letter-spacing:0.15em;"
        "text-transform:uppercase;white-space:nowrap;'>" + text + "</div>"
        "<div style='flex:1;height:1px;background:linear-gradient(270deg,rgba(58,134,255,0.4),transparent);'></div>"
        "</div>"
    )
    st.markdown(html, unsafe_allow_html=True)

def hero():
    html = (
        "<div style='background:linear-gradient(135deg,#0d1b3e 0%,#1a1040 60%,#0d1b3e 100%);"
        "border:1px solid rgba(58,134,255,0.2);border-radius:16px;padding:2.5rem;margin-bottom:2rem;'>"
        "<div style='font-size:0.8rem;color:#3a86ff;letter-spacing:0.2em;text-transform:uppercase;"
        "margin-bottom:0.5rem;font-family:monospace;'>Laboratory Analysis Platform</div>"
        "<div style='font-size:2.8rem;font-weight:800;color:white;margin-bottom:0.5rem;'>Sou"
        "<span style='background:linear-gradient(90deg,#3a86ff,#8338ec);-webkit-background-clip:text;"
        "-webkit-text-fill-color:transparent;background-clip:text;'>Lab</span></div>"
        "<div style='color:rgba(255,255,255,0.5);font-size:1.05rem;max-width:540px;line-height:1.6;"
        "margin-bottom:1.25rem;'>Automated protein assay and peroxisome proliferation analysis. "
        "Paste your raw plate reader data and get publication-ready Excel results instantly.</div>"
        "<div style='display:flex;gap:1.5rem;flex-wrap:wrap;'>"
        "<span style='font-size:0.85rem;color:rgba(255,255,255,0.3);font-family:monospace;'>01 — Set up plate guides</span>"
        "<span style='font-size:0.85rem;color:rgba(255,255,255,0.3);font-family:monospace;'>02 — Paste absorbance data</span>"
        "<span style='font-size:0.85rem;color:rgba(255,255,255,0.3);font-family:monospace;'>03 — Build standard curve</span>"
        "<span style='font-size:0.85rem;color:rgba(255,255,255,0.3);font-family:monospace;'>04 — Enter sample info</span>"
        "<span style='font-size:0.85rem;color:rgba(255,255,255,0.3);font-family:monospace;'>05 — Download results</span>"
        "</div></div>"
    )
    st.markdown(html, unsafe_allow_html=True)

def status_pills(checks):
    html = "<div style='display:flex;flex-wrap:wrap;gap:0.5rem;margin-bottom:1rem;'>"
    for name, ready in checks.items():
        if ready:
            html += (
                "<div style='display:flex;align-items:center;gap:0.4rem;padding:0.28rem 0.75rem;"
                "border-radius:20px;background:rgba(34,197,94,0.1);border:1px solid rgba(34,197,94,0.25);"
                "font-size:0.8rem;color:#166534;'>"
                "<div style='width:6px;height:6px;border-radius:50%;background:#4ade80;flex-shrink:0;'></div>"
                + name + "</div>"
            )
        else:
            html += (
                "<div style='display:flex;align-items:center;gap:0.4rem;padding:0.28rem 0.75rem;"
                "border-radius:20px;background:rgba(251,191,36,0.08);border:1px solid rgba(251,191,36,0.2);"
                "font-size:0.8rem;color:#854d0e;'>"
                "<div style='width:6px;height:6px;border-radius:50%;background:#fbbf24;flex-shrink:0;'></div>"
                + name + "</div>"
            )
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)

def parse_pasted_plate(text):
    if not text or not text.strip():
        return None, "No data pasted."
    try:
        df = pd.read_csv(io.StringIO(text.strip()), sep="\t", header=None)
        if df.shape[0] != 8:
            return None, "Got " + str(df.shape[0]) + " rows — expected 8 (A-H). Do not include row labels."
        if df.shape[1] != 12:
            return None, "Got " + str(df.shape[1]) + " columns — expected 12 (1-12). Do not include column numbers."
        df.index = list("ABCDEFGH")
        df.columns = list(range(1, 13))
        return df, None
    except Exception as e:
        return None, "Could not read data: " + str(e)

def render_plate(plate_key, default_data):
    if plate_key not in st.session_state:
        st.session_state[plate_key] = {}
        for r in rows:
            for c in cols:
                st.session_state[plate_key][f"{r}{c}"] = default_data.get(f"{r}{c}", "")
    header_cols = st.columns([0.5] + [1] * 12)
    header_cols[0].markdown(" ")
    for i, c in enumerate(cols):
        header_cols[i+1].markdown(
            "<p style='text-align:center;font-size:0.8rem;font-weight:700;"
            "color:#3a86ff;margin:0;font-family:monospace;'>" + str(c) + "</p>",
            unsafe_allow_html=True
        )
    for r in rows:
        row_cols = st.columns([0.5] + [1] * 12)
        row_cols[0].markdown(
            "<p style='text-align:center;font-size:0.88rem;font-weight:700;"
            "color:#3a86ff;margin:0;padding-top:7px;font-family:monospace;'>" + r + "</p>",
            unsafe_allow_html=True
        )
        for i, c in enumerate(cols):
            key = f"{r}{c}"
            val = row_cols[i+1].text_input(
                label=key, value=st.session_state[plate_key][key],
                label_visibility="collapsed", key=plate_key + "_well_" + key
            )
            st.session_state[plate_key][key] = val

def validate_plate(plate_key, allow_standards=True):
    standards, samples, errors = {}, {}, []
    for r in rows:
        for c in cols:
            key = f"{r}{c}"
            val = st.session_state[plate_key][key].strip()
            if val == "":
                continue
            elif val.endswith("*"):
                if not allow_standards:
                    errors.append("Well " + key + ": remove the * — no standards on this plate")
                    continue
                try:
                    standards[key] = float(val[:-1])
                except ValueError:
                    errors.append("Well " + key + ": not valid — use a number + asterisk e.g. 2000*")
            else:
                try:
                    samples[key] = int(val)
                except ValueError:
                    errors.append("Well " + key + ": not valid — enter an integer ANID e.g. 101")
    return standards, samples, errors

def show_validation_results(standards, samples, std_key, samp_key):
    if standards:
        st.markdown("**Standards detected**")
        st.dataframe(
            pd.DataFrame([{"Well": k, "Concentration (ug/mL)": v} for k, v in sorted(standards.items())]),
            use_container_width=True
        )
        st.session_state[std_key] = standards
    if samples:
        st.markdown("**Samples detected**")
        st.dataframe(
            pd.DataFrame([{"Well": k, "ANID": v} for k, v in sorted(samples.items())]),
            use_container_width=True
        )
        st.session_state[samp_key] = samples

def build_standard_curve(standards_dict, absorbance_df):
    conc_to_wells = {}
    for well, conc in standards_dict.items():
        conc_to_wells.setdefault(conc, []).append(well)
    rows_data = []
    for conc in sorted(conc_to_wells.keys()):
        abs_values = []
        for well in conc_to_wells[conc]:
            try:
                abs_values.append(float(absorbance_df.loc[well[0], int(well[1:])]))
            except Exception:
                pass
        if not abs_values:
            continue
        avg = np.mean(abs_values)
        rows_data.append({
            "Std (ug/mL)": conc,
            "Absorbance Values": abs_values,
            "Average": round(avg, 4),
            "_is_blank": conc == 0.0
        })
    if not rows_data:
        return None, None, None, None
    summary_df = pd.DataFrame(rows_data)
    x = summary_df["Std (ug/mL)"].values.astype(float)
    y = summary_df["Average"].values.astype(float)
    slope, intercept, r_value, _, _ = stats.linregress(x, y)
    r2 = r_value ** 2
    summary_df["Calc. Conc"] = summary_df["Average"].apply(
        lambda a: round((a - intercept) / slope, 1)
    )
    def pct_err(row):
        if row["_is_blank"] or row["Std (ug/mL)"] == 0:
            return None
        return round(((row["Calc. Conc"] - row["Std (ug/mL)"]) / row["Std (ug/mL)"]) * 100, 1)
    summary_df["%Error"] = summary_df.apply(pct_err, axis=1)
    return slope, intercept, r2, summary_df

def build_sample_table(protein_samples, absorbance_df, slope, intercept, organ_weight_data, organ_rows):
    sex_lookup = {}
    for i in range(organ_rows):
        entry = organ_weight_data[i]
        anid_str = entry["ANID"].strip()
        if anid_str:
            try:
                sex_lookup[int(anid_str)] = entry["Sex"].strip()
            except ValueError:
                pass
    anid_to_wells = {}
    for well, anid in protein_samples.items():
        anid_to_wells.setdefault(anid, []).append(well)
    results = []
    for anid in sorted(anid_to_wells.keys()):
        abs_values = []
        for well in anid_to_wells[anid]:
            try:
                abs_values.append(float(absorbance_df.loc[well[0], int(well[1:])]))
            except Exception:
                pass
        rep1 = abs_values[0] if len(abs_values) > 0 else None
        rep2 = abs_values[1] if len(abs_values) > 1 else None
        avg = np.mean(abs_values) if abs_values else None
        conc_ug = round((avg - intercept) / slope, 4) if avg is not None else None
        conc_mg = round(conc_ug / 1000, 6) if conc_ug is not None else None
        mg_ml_sample = round(conc_mg * 40, 4) if conc_mg is not None else None
        results.append({
            "Sample": anid,
            "Sex": sex_lookup.get(anid, ""),
            "Rep 1": rep1,
            "Rep 2": rep2,
            "Avg": round(avg, 4) if avg is not None else None,
            "1:40 conc (ug/mL)": conc_ug,
            "1:40 conc (mg/mL)": conc_mg,
            "mg/mL sample": mg_ml_sample,
        })
    return pd.DataFrame(results)

def build_perox_results(perox_samples, kinetic_df, time_vals, sample_df, organ_weights_df):
    anid_to_wells = {}
    for well, anid in perox_samples.items():
        anid_to_wells.setdefault(anid, []).append(well)
    protein_lookup = {}
    if sample_df is not None:
        for _, row in sample_df.iterrows():
            protein_lookup[row["Sample"]] = row["mg/mL sample"]
    organ_lookup = {}
    if organ_weights_df is not None:
        for _, row in organ_weights_df.iterrows():
            organ_lookup[row["ANID"]] = row["Weight (g)"] * 1000
    anid_tables, summary_rows = {}, []
    for anid in sorted(anid_to_wells.keys()):
        wells = anid_to_wells[anid]
        rep1_well = wells[0] if len(wells) > 0 else None
        rep2_well = wells[1] if len(wells) > 1 else None
        rep1_vals = kinetic_df[rep1_well].values.astype(float) if rep1_well and rep1_well in kinetic_df.columns else np.full(11, np.nan)
        rep2_vals = kinetic_df[rep2_well].values.astype(float) if rep2_well and rep2_well in kinetic_df.columns else np.full(11, np.nan)
        avg_vals = np.where(
            ~np.isnan(rep1_vals) & ~np.isnan(rep2_vals),
            (rep1_vals + rep2_vals) / 2,
            np.where(~np.isnan(rep1_vals), rep1_vals, rep2_vals)
        )
        valid_mask = ~np.isnan(avg_vals)
        slope_result = None
        if valid_mask.sum() > 1:
            slope_result, _, _, _, _ = stats.linregress(
                np.array(time_vals)[valid_mask], avg_vals[valid_mask]
            )
        table_rows = []
        for i in range(len(time_vals)):
            r1_val = None if np.isnan(rep1_vals[i]) else float(rep1_vals[i])
            r2_val = None if np.isnan(rep2_vals[i]) else float(rep2_vals[i])
            avg_val = None if np.isnan(avg_vals[i]) else round(float(avg_vals[i]), 6)
            table_rows.append({
                "Min": i,
                "Time (min)": time_vals[i],
                "Rep 1 (" + (rep1_well or "") + ")": r1_val,
                "Rep 2 (" + (rep2_well or "") + ")": r2_val,
                "Average": avg_val,
            })
        anid_tables[anid] = pd.DataFrame(table_rows)
        protein_conc = protein_lookup.get(anid)
        organ_weight_mg = organ_lookup.get(anid)
        conc_mM_min = round(slope_result / 91000, 8) if slope_result is not None else None
        convert_nM_min = round(conc_mM_min * 1000000, 4) if conc_mM_min is not None else None
        step_41 = round(protein_conc * 0.2, 6) if protein_conc is not None else None
        step_42 = round(convert_nM_min / step_41, 4) if (convert_nM_min is not None and step_41 not in (None, 0)) else None
        step_43 = round((protein_conc * organ_weight_mg) / 1000, 4) if (protein_conc is not None and organ_weight_mg is not None) else None
        step_44 = round(step_42 * step_43, 4) if (step_42 is not None and step_43 is not None) else None
        summary_rows.append({
            "ANID": anid,
            "Slope": round(slope_result, 8) if slope_result is not None else None,
            "Concentration mM/min": conc_mM_min,
            "Convert mM to nM/min": convert_nM_min,
            " ": "",
            "Protein Concentration (mg/mL)": protein_conc,
            "Organ Weight (mg)": organ_weight_mg,
            "4.1: Total Protein (mg)": step_41,
            "4.2: Specific Activity (nM/min/mg)": step_42,
            "4.3: Scaled Liver Protein (mg)": step_43,
            "4.4: Total Liver Enzyme Activity (nM/min/mg)": step_44,
        })
    return anid_tables, pd.DataFrame(summary_rows)

def style_header(ws, cell, bg="2d3a8c", fg="FFFFFF"):
    cell.font = Font(bold=True, color=fg, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def safe_val(val):
    if isinstance(val, (np.integer,)):
        return int(val)
    elif isinstance(val, (np.floating,)):
        return float(val)
    elif isinstance(val, float) and np.isnan(val):
        return None
    elif val == "":
        return None
    return val

def write_grid_sheet(ws, project_title, plate_state, absorbance_df=None, sheet_label=""):
    ws.cell(row=1, column=1, value=(project_title or "") + " — " + sheet_label).font = Font(bold=True, size=14, color="2d3a8c", name="Arial")
    ws.merge_cells("A1:M1")
    for c in range(1, 13):
        style_header(ws, ws.cell(row=2, column=c+1, value=c), bg="3a86ff")
    for ri, r in enumerate(rows):
        style_header(ws, ws.cell(row=ri+3, column=1, value=r), bg="3a86ff")
        for ci, c in enumerate(cols):
            val = absorbance_df.loc[r, c] if absorbance_df is not None else plate_state.get(f"{r}{c}", "")
            cell = ws.cell(row=ri+3, column=ci+2, value=safe_val(val))
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()
            cell.font = Font(name="Arial")
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 10

def write_standard_curve_sheet(ws, project_title, std_summary_df, slope, intercept, r2):
    ws.cell(row=1, column=1, value=(project_title or "") + " — Standard Curve").font = Font(bold=True, size=14, color="2d3a8c", name="Arial")
    ws.merge_cells("A1:G1")
    for label, val, row_idx in [("Slope", slope, 2), ("Intercept", intercept, 3), ("R2", r2, 4)]:
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True, name="Arial")
        ws.cell(row=row_idx, column=2, value=round(float(val), 8)).font = Font(name="Arial")
    for ci, h in enumerate(["Std (ug/mL)", "Abs 1", "Abs 2", "Average", "Calc. Conc", "%Error"]):
        style_header(ws, ws.cell(row=6, column=ci+1, value=h))
    for ri, row_data in std_summary_df.iterrows():
        abs_vals = row_data["Absorbance Values"]
        for ci, val in enumerate([
            float(row_data["Std (ug/mL)"]),
            float(abs_vals[0]) if len(abs_vals) > 0 else "",
            float(abs_vals[1]) if len(abs_vals) > 1 else "",
            float(row_data["Average"]),
            float(row_data["Calc. Conc"]),
            float(row_data["%Error"]) if row_data["%Error"] is not None else "",
        ]):
            cell = ws.cell(row=ri+7, column=ci+1, value=val)
            cell.font = Font(name="Arial")
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 16

def write_sample_results_sheet(ws, project_title, sample_df):
    ws.cell(row=1, column=1, value=(project_title or "") + " — Sample Results").font = Font(bold=True, size=14, color="2d3a8c", name="Arial")
    ws.merge_cells("A1:H1")
    headers = ["Sample", "Sex", "Rep 1", "Rep 2", "Avg", "1:40 conc (ug/mL)", "1:40 conc (mg/mL)", "mg/mL sample"]
    for ci, h in enumerate(headers):
        style_header(ws, ws.cell(row=2, column=ci+1, value=h))
    for ri, row_data in sample_df.iterrows():
        for ci, h in enumerate(headers):
            cell = ws.cell(row=ri+3, column=ci+1, value=safe_val(row_data[h]))
            cell.font = Font(name="Arial")
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

def build_protein_excel(project_title, protein_plate_state, absorbance_df,
                        std_summary_df, slope, intercept, r2, sample_df):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Protein Plate Guide"
    write_grid_sheet(ws1, project_title, protein_plate_state, sheet_label="Protein Plate Guide")
    write_grid_sheet(wb.create_sheet("Protein Raw Absorbance"), project_title, {},
                     absorbance_df=absorbance_df, sheet_label="Protein Raw Absorbance")
    write_standard_curve_sheet(wb.create_sheet("Standard Curve"), project_title,
                               std_summary_df, slope, intercept, r2)
    write_sample_results_sheet(wb.create_sheet("Sample Results"), project_title, sample_df)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def build_full_excel(project_title, protein_plate_state, protein_absorbance_df,
                     std_summary_df, slope, intercept, r2, sample_df,
                     perox_plate_state, kinetic_df, anid_tables, perox_summary_df):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Protein Plate Guide"
    write_grid_sheet(ws1, project_title, protein_plate_state, sheet_label="Protein Plate Guide")
    write_grid_sheet(wb.create_sheet("Protein Raw Absorbance"), project_title, {},
                     absorbance_df=protein_absorbance_df, sheet_label="Protein Raw Absorbance")
    write_standard_curve_sheet(wb.create_sheet("Standard Curve"), project_title,
                               std_summary_df, slope, intercept, r2)
    write_sample_results_sheet(wb.create_sheet("Sample Results"), project_title, sample_df)
    write_grid_sheet(wb.create_sheet("Peroxisome Plate Guide"), project_title,
                     perox_plate_state, sheet_label="Peroxisome Plate Guide")
    ws6 = wb.create_sheet("Peroxisome Raw Data")
    ws6.cell(row=1, column=1, value=(project_title or "") + " — Peroxisome Raw Kinetic Data").font = Font(bold=True, size=14, color="2d3a8c", name="Arial")
    ws6.merge_cells("A1:" + get_column_letter(97) + "1")
    style_header(ws6, ws6.cell(row=2, column=1, value="Min"), bg="3a86ff")
    for ci, col_name in enumerate(EXPECTED_WELL_COLS):
        style_header(ws6, ws6.cell(row=2, column=ci+2, value=col_name), bg="3a86ff")
    for ri in range(kinetic_df.shape[0]):
        cell = ws6.cell(row=ri+3, column=1, value=int(ri))
        cell.font = Font(bold=True, name="Arial")
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center")
        for ci, col_name in enumerate(EXPECTED_WELL_COLS):
            raw_val = kinetic_df.iloc[ri][col_name]
            cell = ws6.cell(row=ri+3, column=ci+2, value=float(raw_val) if pd.notna(raw_val) else None)
            cell.font = Font(name="Arial")
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
    for col in ws6.columns:
        ws6.column_dimensions[get_column_letter(col[0].column)].width = 10
    ws7 = wb.create_sheet("Peroxisome Results")
    ws7.cell(row=1, column=1, value=(project_title or "") + " — Peroxisome Results").font = Font(bold=True, size=14, color="2d3a8c", name="Arial")
    current_row = 3
    for anid, table_df in anid_tables.items():
        anid_cell = ws7.cell(row=current_row, column=1, value="ANID: " + str(int(anid)))
        anid_cell.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
        anid_cell.fill = PatternFill("solid", start_color="8338ec")
        ws7.merge_cells("A" + str(current_row) + ":" + get_column_letter(len(table_df.columns)) + str(current_row))
        current_row += 1
        for ci, h in enumerate(table_df.columns):
            style_header(ws7, ws7.cell(row=current_row, column=ci+1, value=h))
        current_row += 1
        for _, row_data in table_df.iterrows():
            for ci, h in enumerate(table_df.columns):
                cell = ws7.cell(row=current_row, column=ci+1, value=safe_val(row_data[h]))
                cell.font = Font(name="Arial")
                cell.border = thin_border()
                cell.alignment = Alignment(horizontal="center")
            current_row += 1
        current_row += 2
    current_row += 1
    summary_title = ws7.cell(row=current_row, column=1, value="Summary — All ANIDs")
    summary_title.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    summary_title.fill = PatternFill("solid", start_color="2d3a8c")
    ws7.merge_cells("A" + str(current_row) + ":" + get_column_letter(len(perox_summary_df.columns)) + str(current_row))
    current_row += 1
    for ci, h in enumerate(perox_summary_df.columns):
        style_header(ws7, ws7.cell(row=current_row, column=ci+1, value=h))
    current_row += 1
    for _, row_data in perox_summary_df.iterrows():
        for ci, h in enumerate(perox_summary_df.columns):
            cell = ws7.cell(row=current_row, column=ci+1, value=safe_val(row_data[h]))
            cell.font = Font(name="Arial")
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
        current_row += 1
    for col in ws7.columns:
        ws7.column_dimensions[get_column_letter(col[0].column)].width = 22
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ── UI ────────────────────────────────────────────────────────────────────────

hero()

section_label("Project Setup")

with st.expander("Step 1 — Project Info", expanded=True):
    instr("Give this experiment a name. It will appear as the title on every sheet in your downloaded Excel file.")
    project_title = st.text_input("Project title", placeholder="e.g. Trial 5 — Liver Peroxisome Study")

section_label("Protein Assay")

with st.expander("Step 2 — Protein Assay Plate Guide", expanded=False):
    instr(
        "<b>Standards:</b> enter concentration + asterisk e.g. <code>2000*</code> &nbsp;|&nbsp; "
        "<b>Samples:</b> enter ANID integer e.g. <code>101</code> &nbsp;|&nbsp; "
        "<b>Empty wells:</b> leave blank. "
        "Columns 1 and 2 are pre-filled with the standard duplicate series (A=2000 down to H=0 ug/mL). "
        "Edit any cell freely."
    )
    protein_defaults = {}
    for i, r in enumerate(rows):
        protein_defaults[f"{r}1"] = default_std_series[i]
        protein_defaults[f"{r}2"] = default_std_series[i]
    render_plate("protein_plate", protein_defaults)
    if st.button("Validate Protein Plate", key="val_protein"):
        standards, samples, errors = validate_plate("protein_plate", allow_standards=True)
        if errors:
            st.error("Fix the following:")
            for e in errors:
                st.write("• " + e)
        else:
            st.success("Validated — " + str(len(standards)) + " standard wells and " + str(len(samples)) + " sample wells.")
            show_validation_results(standards, samples, "protein_standards", "protein_samples")

with st.expander("Step 3 — Protein Assay Raw Absorbance Data", expanded=False):
    instr(
        "In Excel, select <b>only</b> the 96 absorbance values — 8 rows x 12 columns. "
        "Do not include the row labels (A-H) or column numbers (1-12). "
        "Press <code>Ctrl+C</code>, click in the box below, press <code>Ctrl+V</code>."
    )
    raw = st.text_area(
        "Paste absorbance data here", height=220,
        placeholder="Select your 8x12 block in Excel, Ctrl+C, click here, Ctrl+V",
        key="protein_absorbance_raw"
    )
    df, error = parse_pasted_plate(raw)
    if raw.strip() and error:
        st.error(error)
    elif df is not None:
        st.success("Received — 8 rows x 12 columns confirmed.")
        st.dataframe(df, use_container_width=True)
        st.session_state["protein_absorbance"] = df

with st.expander("Step 4 — Build Standard Curve", expanded=False):
    instr(
        "Click below to calculate the linear regression from your standards. "
        "The slope and intercept will be used to convert all sample absorbance readings to protein concentrations."
    )
    if st.button("Build Standard Curve", key="build_std"):
        standards = st.session_state.get("protein_standards")
        absorbance = st.session_state.get("protein_absorbance")
        if not standards:
            st.error("Complete Step 2 first — validate your protein plate guide.")
        elif absorbance is None:
            st.error("Complete Step 3 first — paste your protein absorbance data.")
        else:
            slope, intercept, r2, summary_df = build_standard_curve(standards, absorbance)
            if summary_df is None:
                st.error("Could not build curve — check your plate guide and absorbance data.")
            else:
                st.session_state.update({
                    "std_slope": slope, "std_intercept": intercept,
                    "std_r2": r2, "std_summary_df": summary_df
                })
                c1, c2, c3 = st.columns(3)
                c1.metric("Slope", f"{slope:.6f}")
                c2.metric("Intercept", f"{intercept:.6f}")
                c3.metric("R2", f"{r2:.6f}")
                display_rows = []
                for _, row in summary_df.iterrows():
                    abs_vals = row["Absorbance Values"]
                    display_rows.append({
                        "Std (ug/mL)": int(row["Std (ug/mL)"]) if row["Std (ug/mL)"] == int(row["Std (ug/mL)"]) else row["Std (ug/mL)"],
                        "Abs 1": abs_vals[0] if len(abs_vals) > 0 else "",
                        "Abs 2": abs_vals[1] if len(abs_vals) > 1 else "",
                        "Average": row["Average"],
                        "Calc. Conc": row["Calc. Conc"],
                        "%Error": row["%Error"] if row["%Error"] is not None else "",
                    })
                st.dataframe(pd.DataFrame(display_rows), use_container_width=True)

section_label("Peroxisome Proliferation")

with st.expander("Step 5 — Peroxisome Proliferation Plate Guide", expanded=False):
    instr(
        "No standards on this plate — samples only. "
        "Enter the ANID in each well e.g. <code>101</code>. "
        "For duplicates, enter the same ANID in two different wells anywhere on the plate. "
        "The app will automatically pair them and average their readings across all time points."
    )
    render_plate("perox_plate", {})
    if st.button("Validate Peroxisome Plate", key="val_perox"):
        standards, samples, errors = validate_plate("perox_plate", allow_standards=False)
        if errors:
            st.error("Fix the following:")
            for e in errors:
                st.write("• " + e)
        else:
            st.success("Validated — " + str(len(samples)) + " sample wells found.")
            show_validation_results({}, samples, "", "perox_samples")

with st.expander("Step 6 — Peroxisome Kinetic Absorbance Data", expanded=False):
    instr(
        "In your plate reader output you will see columns for Cycle Number, Time, and Temperature on the left, "
        "followed by the well data starting at A1 all the way through H12. "
        "<b>Only paste the well data — starting at A1 and everything that comes after it.</b> "
        "Do not include the Cycle Number, Time, or Temperature columns on the left. "
        "Do not include any well labels (A1, A2...) across the top. "
        "Just the raw absorbance numbers from A1 onward exactly as they appear in the plate reader — nothing outside of that. "
        "Your selection should be exactly <b>11 rows x 96 columns</b>. "
        "Time points will be assigned automatically as minutes 0 through 10."
    )
    perox_kinetic_raw = st.text_area(
        "Paste kinetic data here", height=250,
        placeholder="Select from A1 data onward — 11 rows x 96 columns, no Cycle/Time/Temp columns, no labels — Ctrl+C then Ctrl+V here",
        key="perox_kinetic_raw"
    )
    if perox_kinetic_raw.strip():
        try:
            kinetic_df = pd.read_csv(io.StringIO(perox_kinetic_raw.strip()), sep="\t", header=None)
            if kinetic_df.shape[1] != 96:
                st.error(
                    "Got " + str(kinetic_df.shape[1]) + " columns — expected 96. "
                    "Make sure you have not included the Cycle Number, Time, or Temperature columns. "
                    "Start your selection at the A1 data column only."
                )
            elif kinetic_df.shape[0] != 11:
                st.error("Got " + str(kinetic_df.shape[0]) + " rows — expected 11 (minutes 0-10).")
            else:
                kinetic_df.columns = EXPECTED_WELL_COLS
                kinetic_df.index = list(range(11))
                for w in EXPECTED_WELL_COLS:
                    kinetic_df[w] = pd.to_numeric(kinetic_df[w], errors="coerce")
                st.session_state["perox_time"] = list(range(11))
                st.success("Received — 11 time points x 96 wells confirmed.")
                preview_df = kinetic_df.copy()
                preview_df.index = ["Min " + str(i) for i in range(11)]
                st.dataframe(preview_df, use_container_width=True)
                st.session_state["perox_kinetic"] = kinetic_df
        except Exception as e:
            st.error("Could not read data: " + str(e))

section_label("Sample Metadata")

with st.expander("Step 7 — Organ Weights and Sample Info", expanded=False):
    instr(
        "Enter the ANID, organ weight in grams, and biological sex for each sample. "
        "ANIDs must match exactly what you entered in the plate guides above. "
        "The app converts grams to mg automatically. "
        "Organ weights and sex are required for the full peroxisome proliferation analysis but not for protein-only results. "
        "Use the slider to adjust the number of rows."
    )
    organ_rows = st.slider("Number of samples", min_value=1, max_value=50, value=24)
    if "organ_weight_data" not in st.session_state:
        st.session_state.organ_weight_data = [{"ANID": "", "Weight (g)": "", "Sex": ""} for _ in range(50)]
    while len(st.session_state.organ_weight_data) < 50:
        st.session_state.organ_weight_data.append({"ANID": "", "Weight (g)": "", "Sex": ""})
    for entry in st.session_state.organ_weight_data:
        if "Sex" not in entry:
            entry["Sex"] = ""
    hc1, hc2, hc3 = st.columns(3)
    hc1.markdown("<span style='color:#000000;font-weight:600;'>ANID</span>", unsafe_allow_html=True)
    hc2.markdown("<span style='color:#000000;font-weight:600;'>Organ Weight (g)</span>", unsafe_allow_html=True)
    hc3.markdown("<span style='color:#000000;font-weight:600;'>Sex (M / F)</span>", unsafe_allow_html=True)
    for i in range(organ_rows):
        c1, c2, c3 = st.columns(3)
        with c1:
            v = st.text_input("ANID " + str(i+1), value=st.session_state.organ_weight_data[i]["ANID"],
                              label_visibility="collapsed", key="organ_anid_" + str(i))
            st.session_state.organ_weight_data[i]["ANID"] = v
        with c2:
            v = st.text_input("Weight " + str(i+1), value=st.session_state.organ_weight_data[i]["Weight (g)"],
                              label_visibility="collapsed", key="organ_weight_" + str(i))
            st.session_state.organ_weight_data[i]["Weight (g)"] = v
        with c3:
            v = st.text_input("Sex " + str(i+1), value=st.session_state.organ_weight_data[i]["Sex"],
                              label_visibility="collapsed", key="organ_sex_" + str(i))
            st.session_state.organ_weight_data[i]["Sex"] = v
    if st.button("Validate Sample Info", key="val_organ"):
        organ_entries, organ_errors = [], []
        for i in range(organ_rows):
            entry = st.session_state.organ_weight_data[i]
            anid = entry["ANID"].strip()
            weight = entry["Weight (g)"].strip()
            sex = entry["Sex"].strip().upper()
            if anid == "" and weight == "" and sex == "":
                continue
            if anid == "" or weight == "":
                organ_errors.append("Row " + str(i+1) + ": both ANID and weight are required.")
                continue
            try:
                anid_int = int(anid)
            except ValueError:
                organ_errors.append("Row " + str(i+1) + ": ANID must be an integer.")
                continue
            try:
                weight_float = float(weight)
            except ValueError:
                organ_errors.append("Row " + str(i+1) + ": Weight must be a number.")
                continue
            if sex not in ("M", "F", ""):
                organ_errors.append("Row " + str(i+1) + ": Sex must be M or F.")
                continue
            organ_entries.append({"ANID": anid_int, "Weight (g)": weight_float, "Sex": sex})
        if organ_errors:
            st.error("Fix the following:")
            for e in organ_errors:
                st.write("• " + e)
        else:
            organ_df = pd.DataFrame(organ_entries)
            st.success("Validated — " + str(len(organ_entries)) + " samples.")
            st.dataframe(organ_df, use_container_width=True)
            st.session_state["organ_weights"] = organ_df

with st.expander("Step 8 — ANID Cross Validation", expanded=False):
    instr(
        "Verify every ANID appears in all three inputs before running the full analysis. "
        "Any mismatch is flagged here so you can fix it first. "
        "Only required before running the full analysis — not needed for protein-only results."
    )
    if st.button("Check All ANIDs Match", key="check_anid"):
        protein_samples = st.session_state.get("protein_samples")
        perox_samples = st.session_state.get("perox_samples")
        organ_weights = st.session_state.get("organ_weights")
        missing = []
        if not protein_samples:
            missing.append("Protein plate (Step 2)")
        if not perox_samples:
            missing.append("Peroxisome plate (Step 5)")
        if organ_weights is None:
            missing.append("Organ weights (Step 7)")
        if missing:
            st.error("Complete these first: " + ", ".join(missing))
        else:
            protein_anids = set(protein_samples.values())
            perox_anids = set(perox_samples.values())
            organ_anids = set(organ_weights["ANID"].tolist())
            all_anids = protein_anids | perox_anids | organ_anids
            errors = []
            for anid in sorted(all_anids):
                missing_from = []
                if anid not in protein_anids:
                    missing_from.append("protein plate")
                if anid not in perox_anids:
                    missing_from.append("peroxisome plate")
                if anid not in organ_anids:
                    missing_from.append("organ weights")
                if missing_from:
                    errors.append("ANID " + str(anid) + " missing from: " + ", ".join(missing_from))
            if errors:
                st.error(str(len(errors)) + " mismatch(es) found:")
                for e in errors:
                    st.write("• " + e)
            else:
                st.success("All " + str(len(all_anids)) + " ANIDs match across all inputs.")

section_label("Run Analysis and Export")

with st.expander("Step 9 — Run Analysis and Download Results", expanded=True):

    protein_ready = all([
        st.session_state.get("protein_absorbance") is not None,
        st.session_state.get("std_slope") is not None,
        st.session_state.get("protein_samples") is not None,
    ])

    full_ready = all([
        st.session_state.get("protein_absorbance") is not None,
        st.session_state.get("std_slope") is not None,
        st.session_state.get("protein_samples") is not None,
        st.session_state.get("organ_weights") is not None,
        st.session_state.get("perox_kinetic") is not None,
        st.session_state.get("perox_samples") is not None,
        st.session_state.get("perox_time") is not None,
    ])

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("<span style='color:#000000;font-size:1.2rem;font-weight:700;'>Protein Analysis</span>", unsafe_allow_html=True)
        st.markdown("<span style='color:#000000;font-size:0.85rem;'>Outputs protein plate guide, raw absorbance, standard curve, and sample results. Organ weights not required.</span>", unsafe_allow_html=True)
        st.markdown(" ")
        status_pills({
            "Protein plate validated": st.session_state.get("protein_standards") is not None,
            "Absorbance data pasted": st.session_state.get("protein_absorbance") is not None,
            "Standard curve built": st.session_state.get("std_slope") is not None,
        })
        if protein_ready:
            if st.button("Run Protein Analysis", key="run_protein"):
                with st.spinner("Calculating..."):
                    sample_df = build_sample_table(
                        st.session_state["protein_samples"],
                        st.session_state["protein_absorbance"],
                        st.session_state["std_slope"],
                        st.session_state["std_intercept"],
                        st.session_state.organ_weight_data,
                        organ_rows
                    )
                    st.session_state["sample_df"] = sample_df
                    st.markdown("<span style='color:#000000;font-weight:700;'>Sample Results</span>", unsafe_allow_html=True)
                    st.dataframe(sample_df, use_container_width=True)
                    excel_file = build_protein_excel(
                        project_title,
                        st.session_state["protein_plate"],
                        st.session_state["protein_absorbance"],
                        st.session_state["std_summary_df"],
                        st.session_state["std_slope"],
                        st.session_state["std_intercept"],
                        st.session_state["std_r2"],
                        sample_df
                    )
                    filename = (project_title.replace(" ", "_") if project_title else "results") + "_protein.xlsx"
                    st.download_button(
                        "Download Protein Results (.xlsx)", data=excel_file,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_protein"
                    )
        else:
            st.info("Complete Steps 2, 3, and 4 to enable protein analysis.")

    with col2:
        st.markdown("<span style='color:#000000;font-size:1.2rem;font-weight:700;'>Full Analysis</span>", unsafe_allow_html=True)
        st.markdown("<span style='color:#000000;font-size:0.85rem;'>Includes protein results plus peroxisome kinetic tables and all four calculation steps. Requires organ weights.</span>", unsafe_allow_html=True)
        st.markdown(" ")
        status_pills({
            "Protein plate validated": st.session_state.get("protein_standards") is not None,
            "Absorbance data pasted": st.session_state.get("protein_absorbance") is not None,
            "Standard curve built": st.session_state.get("std_slope") is not None,
            "Organ weights validated": st.session_state.get("organ_weights") is not None,
            "Peroxisome plate validated": st.session_state.get("perox_samples") is not None,
            "Kinetic data pasted": st.session_state.get("perox_kinetic") is not None,
        })
        if full_ready:
            if st.button("Run Full Analysis", key="run_full"):
                with st.spinner("Calculating..."):
                    sample_df = st.session_state.get("sample_df")
                    if sample_df is None:
                        sample_df = build_sample_table(
                            st.session_state["protein_samples"],
                            st.session_state["protein_absorbance"],
                            st.session_state["std_slope"],
                            st.session_state["std_intercept"],
                            st.session_state.organ_weight_data,
                            organ_rows
                        )
                        st.session_state["sample_df"] = sample_df
                    anid_tables, perox_summary_df = build_perox_results(
                        st.session_state["perox_samples"],
                        st.session_state["perox_kinetic"],
                        st.session_state["perox_time"],
                        sample_df,
                        st.session_state["organ_weights"]
                    )
                    st.markdown("<span style='color:#000000;font-weight:700;'>Peroxisome Summary</span>", unsafe_allow_html=True)
                    st.dataframe(perox_summary_df, use_container_width=True)
                    excel_file = build_full_excel(
                        project_title,
                        st.session_state["protein_plate"],
                        st.session_state["protein_absorbance"],
                        st.session_state["std_summary_df"],
                        st.session_state["std_slope"],
                        st.session_state["std_intercept"],
                        st.session_state["std_r2"],
                        sample_df,
                        st.session_state["perox_plate"],
                        st.session_state["perox_kinetic"],
                        anid_tables,
                        perox_summary_df
                    )
                    filename = (project_title.replace(" ", "_") if project_title else "results") + "_full.xlsx"
                    st.download_button(
                        "Download Full Results (.xlsx)", data=excel_file,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_full"
                    )
        else:
            st.info("Complete all Steps 2-7 to enable full analysis.")

st.markdown("<br>", unsafe_allow_html=True)
st.markdown("<span style='color:rgba(255,255,255,0.3);font-size:0.8rem;'>SouLab — Plate Reader Analysis Platform</span>", unsafe_allow_html=True)
